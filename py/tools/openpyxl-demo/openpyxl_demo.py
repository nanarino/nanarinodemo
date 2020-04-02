from pprint import pprint


from openpyxl import load_workbook


def read_excel_xlsx(path: str,
                    sheet_index: int = 0,
                    is_read_head: bool = False):
    ''' @desc 只读电子表格函数
    @param path: xlsx的文件路径
    @param sheet_index: table的序号(0即Sheet1) 
        自定义的table命名可能无法用下标找到导致抛错
    @param is_read_head: 是否读取表头(即第一行)
    @return -> 返回列表嵌套元组的结构(和数据库查询的格式结果一致)'''
    wb = load_workbook(path, read_only=True)
    table = wb[wb.sheetnames[sheet_index]]
    #最大行
    rows = table.max_row
    #最大列
    cols = table.max_column
    if rows < 2:
        raise Exception("读取不到数据,或者只有一个表头,可以尝试把table命名为\'Sheet%d\'" %
                        (1 + sheet_index))
    #单元格读取
    #print(table.cell(row = 1, column = 1).value)

    #逐行读取 int(是否读表头（第一行）) +1 默认不读
    querylist = map(
        lambda row: tuple(map(lambda col: str(col.value).strip(), row)),
        table.iter_rows(min_row=int(not is_read_head) + 1,
                        max_col=cols,
                        max_row=rows))
    #pprint(list(querylist), width=60, compact=True)
    return querylist


try:
    querylist = read_excel_xlsx('./r.xlsx', 0)  #, is_read_head = True)
except Exception as e:
    input(e)
    exit()


#取出原有表头
#head = list(next(querylist))
#自定义表头
head = ['机构名称', '优惠卡编号', '姓名', '性别', '身份证号', '出生日期', '手机号', '通讯地址', '积分']
#表头插入到数据中
querylist.insert(0, head)


from openpyxl import Workbook


def write_excel_xlsx(path: str, sheet_index: int, value: list):
    ''' @desc 新写入电子表格函数
    @param path: 要生成xlsx文件的路径
    @param value: 要写入的数据'''
    index = len(value)
    wb = Workbook()
    #sheet = wb.active 等价于下一行传0,即"Sheet1" 也可以用sheet.title自定义
    sheet = wb[wb.sheetnames[sheet_index]]
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=value[i][j])
    wb.save(path)


try:
    write_excel_xlsx('./w.xlsx', 0, querylist)
except Exception as e:
    input(e)
    exit()

input("执行完毕")